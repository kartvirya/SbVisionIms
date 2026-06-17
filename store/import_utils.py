"""Bulk import helpers for inventory, customers, suppliers, purchases, and sales."""

import csv
import io
from decimal import Decimal, InvalidOperation

from openpyxl import Workbook, load_workbook

from accounts.models import Brand, Customer, Logistics, Vendor
from bills.models import Bill
from invoice.models import Invoice
from store.models import Category, Delivery, Item
from transactions.models import Purchase, PurchaseLine, Sale, SaleDetail
from transactions.services import reconcile_ledger_stock_to_target, sync_item_quantity_cache


INVENTORY_HEADERS = [
    "vendor",
    "brand",
    "category",
    "name",
    "hs_code",
    "sku",
    "description",
    "quantity",
    "cost_price",
    "price",
    "low_stock_threshold",
]

CUSTOMER_HEADERS = [
    "first_name",
    "last_name",
    "phone",
    "email",
    "address",
    "opening_balance",
]

SUPPLIER_HEADERS = [
    "name",
    "phone_number",
    "pan_number",
    "vat_number",
    "address",
    "opening_balance",
    "brands",
]

PURCHASE_HEADERS = [
    "vendor",
    "bill_number",
    "order_date",
    "receipt_status",
    "item_name",
    "quantity",
    "unit_price",
    "discount_amount",
    "vat_percentage",
    "amount_paid",
    "description",
]

SALE_HEADERS = [
    "sale_reference",
    "sale_date",
    "customer_first_name",
    "customer_last_name",
    "customer_phone",
    "vendor",
    "brand",
    "category",
    "item_name",
    "quantity",
    "stock",
    "unit_price",
    "tax_percentage",
    "amount_paid",
    "description",
]

IMPORT_REQUIRED_HEADERS = {
    "sales": ["customer_first_name", "item_name"],
}

DELIVERY_HEADERS = [
    "item_name",
    "customer_name",
    "phone_number",
    "location",
    "date",
    "logistics",
    "tracking_number",
    "is_delivered",
]

INVOICE_HEADERS = [
    "customer_name",
    "contact_number",
    "item_name",
    "price_per_item",
    "quantity",
    "shipping",
]

BILL_HEADERS = [
    "institution_name",
    "phone_number",
    "email",
    "address",
    "description",
    "payment_details",
    "amount",
    "status",
]


def inventory_template_workbook():
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventory"
    ws.append(INVENTORY_HEADERS)
    ws.append(
        [
            "ABC Traders",
            "Samsung",
            "General",
            "Sample Product",
            "8471.30",
            "SKU-001",
            "Optional description",
            10,
            50,
            100,
            5,
        ]
    )
    return wb


def customer_template_workbook():
    wb = Workbook()
    ws = wb.active
    ws.title = "Customers"
    ws.append(CUSTOMER_HEADERS)
    ws.append(["Ram", "Sharma", "9800000000", "ram@example.com", "Kathmandu", 0])
    return wb


def supplier_template_workbook():
    wb = Workbook()
    ws = wb.active
    ws.title = "Suppliers"
    ws.append(SUPPLIER_HEADERS)
    ws.append(
        [
            "ABC Traders",
            "01-4212345",
            "123456789",
            "601234567",
            "Kathmandu",
            0,
            "Samsung; LG",
        ]
    )
    return wb


def purchase_template_workbook():
    wb = Workbook()
    ws = wb.active
    ws.title = "Purchases"
    ws.append(PURCHASE_HEADERS)
    ws.append(
        [
            "ABC Traders",
            "BILL-100",
            "2026-06-01",
            "S",
            "Sample Product",
            5,
            50,
            0,
            13,
            0,
            "Imported purchase",
        ]
    )
    return wb


def sale_template_workbook():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sales"
    ws.append(SALE_HEADERS)
    ws.append(
        [
            "SALE-001",
            "2026-06-01",
            "Ram",
            "Sharma",
            "9800000000",
            "ABC Traders",
            "Samsung",
            "Electronics",
            "Sample Product",
            1,
            5,
            100,
            13,
            113,
            "Delete this sample row before importing your data",
        ]
    )
    return wb


def delivery_template_workbook():
    wb = Workbook()
    ws = wb.active
    ws.title = "Deliveries"
    ws.append(DELIVERY_HEADERS)
    ws.append(
        [
            "Sample Product",
            "Ram Sharma",
            "9800000000",
            "Kathmandu",
            "2026-06-01",
            "",
            "TRK-001",
            "no",
        ]
    )
    return wb


def invoice_template_workbook():
    wb = Workbook()
    ws = wb.active
    ws.title = "Invoices"
    ws.append(INVOICE_HEADERS)
    ws.append(
        [
            "Ram Sharma",
            "9800000000",
            "Sample Product",
            100,
            2,
            50,
        ]
    )
    return wb


def bill_template_workbook():
    wb = Workbook()
    ws = wb.active
    ws.title = "Bills"
    ws.append(BILL_HEADERS)
    ws.append(
        [
            "Electricity Board",
            "014412345",
            "info@example.com",
            "Kathmandu",
            "Monthly bill",
            "Bank transfer",
            1500,
            "no",
        ]
    )
    return wb


TEMPLATE_BUILDERS = {
    "inventory": inventory_template_workbook,
    "customers": customer_template_workbook,
    "suppliers": supplier_template_workbook,
    "purchases": purchase_template_workbook,
    "sales": sale_template_workbook,
    "deliveries": delivery_template_workbook,
    "invoices": invoice_template_workbook,
    "bills": bill_template_workbook,
}

IMPORT_HEADERS = {
    "inventory": INVENTORY_HEADERS,
    "customers": CUSTOMER_HEADERS,
    "suppliers": SUPPLIER_HEADERS,
    "purchases": PURCHASE_HEADERS,
    "sales": SALE_HEADERS,
    "deliveries": DELIVERY_HEADERS,
    "invoices": INVOICE_HEADERS,
    "bills": BILL_HEADERS,
}


def _cell_str(value):
    if value is None:
        return ""
    return str(value).strip()


def _phone_str(value):
    """Preserve phone formatting (leading zeros, dashes) from Excel/CSV cells."""
    if value is None:
        return None
    if isinstance(value, float) and value.is_integer():
        raw = str(int(value))
    elif isinstance(value, int):
        raw = str(value)
    else:
        raw = _cell_str(value)
    if not raw:
        return None
    if raw.endswith(".0") and raw[:-2].isdigit():
        raw = raw[:-2]
    return raw


def _split_list_cell(value):
    """Split a cell into multiple values (semicolon or comma separated)."""
    raw = _cell_str(value)
    if not raw:
        return []
    separator = ";" if ";" in raw else ","
    return [part.strip() for part in raw.split(separator) if part.strip()]


def _parse_decimal(value, default="0"):
    try:
        return Decimal(str(value if value not in (None, "") else default))
    except (InvalidOperation, ValueError):
        return Decimal(default)


def _parse_bool(value):
    text = _cell_str(value).lower()
    if text in ("1", "true", "yes", "y", "paid"):
        return True
    if text in ("0", "false", "no", "n", "unpaid", ""):
        return False
    return False


def _parse_import_datetime(value):
    from datetime import datetime

    from django.utils import timezone
    from django.utils.dateparse import parse_date, parse_datetime

    raw = _cell_str(value)
    if not raw:
        return timezone.now()
    parsed = parse_datetime(raw)
    if not parsed:
        d = parse_date(raw)
        if d:
            parsed = timezone.make_aware(datetime.combine(d, datetime.min.time()))
    return parsed or timezone.now()


def _read_csv_rows(uploaded_file, expected_headers, required_headers=None):
    required_headers = required_headers or expected_headers
    text = uploaded_file.read().decode("utf-8-sig")
    reader = csv.reader(io.StringIO(text))
    header_row = next(reader, None)
    if not header_row:
        return [], ["Empty CSV file."]
    headers = [_cell_str(h).lower().replace(" ", "_") for h in header_row]
    missing = [h for h in required_headers if h not in headers]
    if missing:
        return [], [f"Missing columns: {', '.join(missing)}"]
    known = set(expected_headers)
    idx = {h: headers.index(h) for h in headers if h in known}
    out = []
    for line in reader:
        if not any(line):
            continue
        out.append(
            {
                h: line[idx[h]] if h in idx and idx[h] < len(line) else None
                for h in expected_headers
            }
        )
    return out, []


def _read_xlsx_rows(uploaded_file, expected_headers, required_headers=None):
    required_headers = required_headers or expected_headers
    wb = load_workbook(uploaded_file, read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    header_row = next(rows_iter, None)
    if not header_row:
        return [], ["Empty spreadsheet."]
    headers = [_cell_str(h).lower().replace(" ", "_") for h in header_row]
    missing = [h for h in required_headers if h not in headers]
    if missing:
        return [], [f"Missing columns: {', '.join(missing)}"]
    known = set(expected_headers)
    idx = {h: headers.index(h) for h in headers if h in known}
    out = []
    for line in rows_iter:
        if not any(line):
            continue
        out.append(
            {
                h: line[idx[h]] if h in idx and idx[h] < len(line) else None
                for h in expected_headers
            }
        )
    return out, []


def read_sheet_rows(uploaded_file, expected_headers, required_headers=None):
    name = (getattr(uploaded_file, "name", "") or "").lower()
    if name.endswith(".csv"):
        return _read_csv_rows(uploaded_file, expected_headers, required_headers)
    return _read_xlsx_rows(uploaded_file, expected_headers, required_headers)


def import_inventory_rows(rows):
    created = updated = 0
    errors = []
    for i, row in enumerate(rows, start=2):
        name = _cell_str(row.get("name"))
        if not name:
            errors.append(f"Row {i}: name is required.")
            continue
        cat_name = _cell_str(row.get("category")) or "General"
        category, _ = Category.objects.get_or_create(name=cat_name)
        vendor = None
        brand = None
        vendor_name = _cell_str(row.get("vendor"))
        if vendor_name:
            vendor, _ = Vendor.objects.get_or_create(name=vendor_name)
            brand_name = _cell_str(row.get("brand"))
            if brand_name:
                brand, _ = Brand.objects.get_or_create(
                    vendor=vendor,
                    name=brand_name,
                )
        target_qty = int(_parse_decimal(row.get("quantity"), 0))
        sku = _cell_str(row.get("sku"))
        hs_code = _cell_str(row.get("hs_code"))
        defaults = {
            "description": _cell_str(row.get("description")) or name,
            "category": category,
            "cost_price": float(_parse_decimal(row.get("cost_price"), 0)),
            "price": float(_parse_decimal(row.get("price"), 0)),
            "vendor": vendor,
            "brand": brand,
            "hs_code": hs_code,
            "low_stock_threshold": int(_parse_decimal(row.get("low_stock_threshold"), 10)),
            "quantity": 0,
        }
        if sku:
            defaults["sku"] = sku
        lookup = {"sku": sku} if sku else {"name": name}
        item, was_created = Item.objects.update_or_create(defaults=defaults, **lookup)
        reconcile_ledger_stock_to_target(
            item, target_qty, notes=f"Excel import row {i}"
        )
        sync_item_quantity_cache([item])
        if was_created:
            created += 1
        else:
            updated += 1
    return created, updated, errors


def import_customer_rows(rows):
    from accounts.customer_utils import (
        find_customer_by_identity,
        get_or_create_customer_by_identity,
    )

    created = updated = 0
    errors = []
    for i, row in enumerate(rows, start=2):
        first = _cell_str(row.get("first_name"))
        if not first:
            errors.append(f"Row {i}: first_name is required.")
            continue
        last = _cell_str(row.get("last_name")) or None
        phone = _phone_str(row.get("phone"))
        opening = row.get("opening_balance")
        defaults = {
            "email": _cell_str(row.get("email")) or None,
            "address": _cell_str(row.get("address")) or None,
            "opening_balance": _parse_decimal(opening, 0)
            if opening not in (None, "")
            else 0,
        }
        existing = find_customer_by_identity(
            first_name=first,
            last_name=last,
            phone=phone,
        )
        if existing:
            existing.first_name = first
            existing.last_name = last
            if phone:
                existing.phone = phone
            if defaults["email"]:
                existing.email = defaults["email"]
            if defaults["address"]:
                existing.address = defaults["address"]
            if opening not in (None, ""):
                existing.opening_balance = defaults["opening_balance"]
            existing.save()
            updated += 1
            continue
        get_or_create_customer_by_identity(
            first_name=first,
            last_name=last,
            phone=phone,
            defaults=defaults,
        )
        created += 1
    return created, updated, errors


def import_supplier_rows(rows):
    created = updated = 0
    errors = []
    for i, row in enumerate(rows, start=2):
        name = _cell_str(row.get("name"))
        if not name:
            errors.append(f"Row {i}: name is required.")
            continue
        defaults = {
            "phone_number": _phone_str(row.get("phone_number")),
            "address": _cell_str(row.get("address")) or None,
        }
        pan_number = _cell_str(row.get("pan_number"))
        if pan_number:
            defaults["pan_number"] = pan_number
        vat_number = _cell_str(row.get("vat_number"))
        if vat_number:
            defaults["vat_number"] = vat_number
        opening = row.get("opening_balance")
        if opening not in (None, ""):
            defaults["opening_balance"] = _parse_decimal(opening, 0)
        vendor, was_created = Vendor.objects.update_or_create(
            name=name,
            defaults=defaults,
        )
        for brand_name in _split_list_cell(row.get("brands")):
            Brand.objects.get_or_create(vendor=vendor, name=brand_name)
        if was_created:
            created += 1
        else:
            updated += 1
    return created, updated, errors


def import_purchase_rows(rows):
    from datetime import datetime

    from django.utils.dateparse import parse_datetime, parse_date
    from django.utils import timezone

    created = updated = 0
    errors = []
    for group in _group_purchase_import_rows(rows):
        first_row = group[0][1]
        vendor_name = _cell_str(first_row.get("vendor"))
        if not vendor_name:
            errors.append(f"Row {group[0][0] + 2}: vendor is required.")
            continue

        line_specs = []
        group_errors = []
        for row_index, row in group:
            item_name = _cell_str(row.get("item_name"))
            if not item_name:
                group_errors.append(f"Row {row_index + 2}: item_name is required.")
                continue
            line_specs.append((row_index, row, item_name))

        if group_errors:
            errors.extend(group_errors)
            continue
        if not line_specs:
            continue

        try:
            vendor, _ = Vendor.objects.get_or_create(name=vendor_name)
            bill_number = _cell_str(first_row.get("bill_number"))
            if not bill_number:
                bill_number = f"import-{group[0][0] + 2}"

            purchase = Purchase.objects.filter(
                vendor=vendor,
                bill_number=bill_number,
            ).first()
            was_new = purchase is None

            if was_new:
                order_raw = _cell_str(first_row.get("order_date"))
                order_date = timezone.now()
                if order_raw:
                    parsed = parse_datetime(order_raw)
                    if not parsed:
                        d = parse_date(order_raw)
                        if d:
                            parsed = timezone.make_aware(
                                datetime.combine(d, datetime.min.time())
                            )
                    if parsed:
                        order_date = parsed
                receipt_status = (
                    _cell_str(first_row.get("receipt_status")) or "P"
                ).upper()[:1]
                if receipt_status not in ("P", "S"):
                    receipt_status = "P"
                purchase = Purchase.objects.create(
                    vendor=vendor,
                    bill_number=bill_number,
                    order_date=order_date,
                    receipt_status=receipt_status,
                    discount_amount=_parse_decimal(
                        first_row.get("discount_amount"), 0
                    ),
                    vat_percentage=float(
                        _parse_decimal(first_row.get("vat_percentage"), 0)
                    ),
                    amount_paid=_parse_decimal(first_row.get("amount_paid"), 0),
                    description=_cell_str(first_row.get("description")),
                )
                created += 1
            else:
                updated += 1

            for row_index, row, item_name in line_specs:
                category, _ = Category.objects.get_or_create(name="General")
                unit_price = _parse_decimal(row.get("unit_price"), 0)
                item, _ = Item.objects.get_or_create(
                    name=item_name,
                    defaults={
                        "description": item_name,
                        "category": category,
                        "vendor": vendor,
                        "cost_price": float(unit_price),
                        "price": float(unit_price * Decimal("1.3"))
                        if unit_price
                        else 0,
                        "quantity": 0,
                    },
                )
                qty = int(_parse_decimal(row.get("quantity"), 1))
                line, line_created = PurchaseLine.objects.get_or_create(
                    purchase=purchase,
                    item=item,
                    defaults={
                        "quantity": max(qty, 1),
                        "unit_price": unit_price,
                    },
                )
                if not line_created:
                    line.quantity = max(qty, 1)
                    line.unit_price = unit_price
                    line.save(update_fields=["quantity", "unit_price"])

            purchase.save()
            from transactions.services import sync_purchase_inventory_transaction

            sync_purchase_inventory_transaction(purchase=purchase)
        except Exception as exc:
            errors.append(f"Row {group[0][0] + 2}: {exc}")
    return created, updated, errors


def _resolve_customer_from_row(row):
    from accounts.customer_utils import get_or_create_customer_by_identity

    first = _cell_str(row.get("customer_first_name"))
    last = _cell_str(row.get("customer_last_name")) or None
    phone = _phone_str(row.get("customer_phone"))
    customer, _ = get_or_create_customer_by_identity(
        first_name=first,
        last_name=last,
        phone=phone,
    )
    return customer


def _resolve_item_from_sale_row(row):
    item_name = _cell_str(row.get("item_name"))
    category_name = _cell_str(row.get("category")) or "General"
    category, _ = Category.objects.get_or_create(name=category_name)
    vendor = None
    brand = None
    vendor_name = _cell_str(row.get("vendor"))
    if vendor_name:
        vendor, _ = Vendor.objects.get_or_create(name=vendor_name)
        brand_name = _cell_str(row.get("brand"))
        if brand_name:
            brand, _ = Brand.objects.get_or_create(
                vendor=vendor,
                name=brand_name,
                defaults={"is_active": True},
            )
    unit_price = _parse_decimal(row.get("unit_price"), 0)
    item = Item.objects.filter(name=item_name).first()
    if item:
        updates = []
        if item.category_id != category.id:
            item.category = category
            updates.append("category")
        if vendor and item.vendor_id != vendor.id:
            item.vendor = vendor
            updates.append("vendor")
        if brand and item.brand_id != brand.id:
            item.brand = brand
            updates.append("brand")
        if unit_price and not item.price:
            item.price = float(unit_price)
            updates.append("price")
        if updates:
            item.save(update_fields=updates)
    else:
        item = Item.objects.create(
            name=item_name,
            description=_cell_str(row.get("description")) or item_name,
            category=category,
            vendor=vendor,
            brand=brand,
            quantity=0,
            price=float(unit_price),
        )
    return item, unit_price


def _ensure_sale_stock(item, qty, stock_cell, row_num):
    """Set ledger stock so a sale of qty can post; stock column = on-hand after sale."""
    from store.stock_utils import get_sellable_stock

    qty_int = max(int(qty), 1)
    stock_raw = stock_cell
    if stock_raw not in (None, ""):
        stock_after = int(_parse_decimal(stock_raw, 0))
        target = max(stock_after + qty_int, 0)
    else:
        available = int(get_sellable_stock(item))
        target = available if available >= qty_int else qty_int
    reconcile_ledger_stock_to_target(
        item,
        target,
        notes=f"Pre-import stock row {row_num}",
    )
    sync_item_quantity_cache([item])


def _group_import_rows(rows, key_fn, inherit_blank=True):
    """
    Group consecutive spreadsheet rows into one bill/sale.
    Blank keys inherit the previous group's key when inherit_blank is True.
    """
    groups = []
    current = []
    active_key = None

    for index, row in enumerate(rows):
        key = key_fn(row)
        if not key and inherit_blank and active_key is not None:
            current.append((index, row))
            continue
        if not key:
            if current:
                groups.append(current)
                current = []
            active_key = None
            groups.append([(index, row)])
            continue
        if key == active_key:
            current.append((index, row))
        else:
            if current:
                groups.append(current)
            active_key = key
            current = [(index, row)]

    if current:
        groups.append(current)
    return groups


def _sale_date_group_key(row):
    raw = _cell_str(row.get("sale_date"))
    if not raw:
        return "nodate"
    from datetime import datetime

    from django.utils.dateparse import parse_date, parse_datetime

    parsed = parse_datetime(raw)
    if parsed:
        return parsed.date().isoformat()
    day = parse_date(raw)
    if day:
        return day.isoformat()
    return raw


def _customer_group_key(row):
    from accounts.customer_utils import customer_identity_key

    return customer_identity_key(
        _cell_str(row.get("customer_first_name")),
        _cell_str(row.get("customer_last_name")) or None,
        _phone_str(row.get("customer_phone")),
    )


def _sale_import_key(row):
    ref = _cell_str(row.get("sale_reference"))
    if ref:
        return f"ref::{ref}"
    return f"sale::{_customer_group_key(row)}::{_sale_date_group_key(row)}"


def _purchase_import_key(row):
    vendor = _cell_str(row.get("vendor"))
    bill = _cell_str(row.get("bill_number"))
    if bill:
        return f"bill::{vendor}::{bill}"
    if vendor:
        return f"vendor::{vendor}"
    return ""


def _group_sale_import_rows(rows):
    return _group_import_rows(rows, _sale_import_key, inherit_blank=True)


def _group_purchase_import_rows(rows):
    return _group_import_rows(rows, _purchase_import_key, inherit_blank=True)


def import_sale_rows(rows):
    from django.db import transaction as db_transaction

    from transactions.models import CustomerPayment
    from transactions.services import create_sale_transaction

    created = 0
    errors = []
    for group in _group_sale_import_rows(rows):
        first_row = group[0][1]
        first_name = _cell_str(first_row.get("customer_first_name"))
        if not first_name:
            errors.append(f"Row {group[0][0] + 2}: customer_first_name is required.")
            continue

        line_items = []
        group_errors = []
        for row_index, row in group:
            item_name = _cell_str(row.get("item_name"))
            if not item_name:
                group_errors.append(f"Row {row_index + 2}: item_name is required.")
                continue
            try:
                item, unit_price = _resolve_item_from_sale_row(row)
                qty = _parse_decimal(row.get("quantity"), 1)
                if qty <= 0:
                    qty = Decimal("1")
                _ensure_sale_stock(item, qty, row.get("stock"), row_index + 2)
                line_items.append(
                    {
                        "row_index": row_index + 2,
                        "item": item,
                        "unit_price": unit_price,
                        "quantity": qty,
                        "description": _cell_str(row.get("description")),
                    }
                )
            except Exception as exc:
                group_errors.append(f"Row {row_index + 2}: {exc}")

        if group_errors:
            errors.extend(group_errors)
            continue
        if not line_items:
            continue

        try:
            customer = _resolve_customer_from_row(first_row)
            import_ref = _sale_import_key(first_row)
            if import_ref and Sale.objects.filter(import_reference=import_ref).exists():
                existing = Sale.objects.filter(import_reference=import_ref).first()
                errors.append(
                    f"Row {group[0][0] + 2}: sale_reference '{import_ref}' "
                    f"was already imported (sale #{existing.id})."
                )
                continue

            sale_date = _parse_import_datetime(first_row.get("sale_date"))
            tax_pct = _parse_decimal(first_row.get("tax_percentage"), 13)
            sub_total = sum(
                (line["unit_price"] * line["quantity"] for line in line_items),
                Decimal("0"),
            )
            tax_amount = (sub_total * (tax_pct / Decimal("100"))).quantize(
                Decimal("0.01")
            )
            grand_total = sub_total + tax_amount

            amount_paid = Decimal("0")
            for _, row in group:
                row_paid = _parse_decimal(row.get("amount_paid"), 0)
                if row_paid > 0:
                    amount_paid = row_paid
                    break
            if amount_paid == 0:
                amount_paid = grand_total
            elif amount_paid < grand_total and len(line_items) == 1:
                line = line_items[0]
                unit_with_tax = (
                    line["unit_price"]
                    * (Decimal("1") + tax_pct / Decimal("100"))
                ).quantize(Decimal("0.01"))
                if amount_paid == unit_with_tax and line["quantity"] > 1:
                    amount_paid = grand_total

            notes = next(
                (line["description"] for line in line_items if line["description"]),
                "",
            )
            with db_transaction.atomic():
                sale = Sale.objects.create(
                    customer=customer,
                    sub_total=sub_total,
                    grand_total=grand_total,
                    tax_amount=tax_amount,
                    tax_percentage=float(tax_pct),
                    amount_paid=Decimal("0"),
                    amount_change=Decimal("0"),
                    date_added=sale_date,
                    import_reference=import_ref,
                )
                for line in line_items:
                    line_sub = line["unit_price"] * line["quantity"]
                    SaleDetail.objects.create(
                        sale=sale,
                        item=line["item"],
                        price=float(line["unit_price"]),
                        quantity=line["quantity"],
                        total_detail=float(line_sub),
                    )
                inventory_transaction = create_sale_transaction(
                    customer=customer,
                    items=[
                        {
                            "item": line["item"].id,
                            "quantity": line["quantity"],
                            "unit_price": line["unit_price"],
                        }
                        for line in line_items
                    ],
                    notes=notes or f"Sale #{sale.id}",
                )
                sale.inventory_transaction = inventory_transaction
                sale.save(update_fields=["inventory_transaction"])
                inventory_transaction.source_ref = f"sale:{sale.id}"
                inventory_transaction.save(update_fields=["source_ref"])
                if amount_paid > 0:
                    CustomerPayment.objects.create(
                        sale=sale,
                        amount=amount_paid,
                        method="cash",
                        notes="Import",
                    )
                else:
                    sale.save()
            created += 1
        except Exception as exc:
            errors.append(f"Row {group[0][0] + 2}: {exc}")
    return created, 0, errors


def import_delivery_rows(rows):
    created = 0
    errors = []
    for i, row in enumerate(rows, start=2):
        item_name = _cell_str(row.get("item_name"))
        customer_name = _cell_str(row.get("customer_name"))
        if not item_name or not customer_name:
            errors.append(f"Row {i}: item_name and customer_name are required.")
            continue
        item = Item.objects.filter(name=item_name).first()
        if not item:
            errors.append(f"Row {i}: item '{item_name}' not found.")
            continue
        logistics = None
        logistics_name = _cell_str(row.get("logistics"))
        if logistics_name:
            logistics = Logistics.objects.filter(name=logistics_name).first()
            if not logistics:
                errors.append(f"Row {i}: logistics '{logistics_name}' not found.")
                continue
        try:
            Delivery.objects.create(
                item=item,
                customer_name=customer_name,
                phone_number=_cell_str(row.get("phone_number")) or None,
                location=_cell_str(row.get("location")) or None,
                date=_parse_import_datetime(row.get("date")),
                logistics=logistics,
                tracking_number=_cell_str(row.get("tracking_number")) or None,
                is_delivered=_parse_bool(row.get("is_delivered")),
            )
            created += 1
        except Exception as exc:
            errors.append(f"Row {i}: {exc}")
    return created, 0, errors


def import_invoice_rows(rows):
    created = 0
    errors = []
    for i, row in enumerate(rows, start=2):
        customer_name = _cell_str(row.get("customer_name"))
        contact_number = _cell_str(row.get("contact_number"))
        item_name = _cell_str(row.get("item_name"))
        if not customer_name or not contact_number or not item_name:
            errors.append(
                f"Row {i}: customer_name, contact_number, and item_name are required."
            )
            continue
        item = Item.objects.filter(name=item_name).first()
        if not item:
            errors.append(f"Row {i}: item '{item_name}' not found.")
            continue
        try:
            Invoice.objects.create(
                customer_name=customer_name,
                contact_number=contact_number,
                item=item,
                price_per_item=float(_parse_decimal(row.get("price_per_item"), 0)),
                quantity=float(_parse_decimal(row.get("quantity"), 1)),
                shipping=float(_parse_decimal(row.get("shipping"), 0)),
            )
            created += 1
        except Exception as exc:
            errors.append(f"Row {i}: {exc}")
    return created, 0, errors


def import_bill_rows(rows):
    created = 0
    errors = []
    for i, row in enumerate(rows, start=2):
        institution_name = _cell_str(row.get("institution_name"))
        payment_details = _cell_str(row.get("payment_details"))
        if not institution_name or not payment_details:
            errors.append(
                f"Row {i}: institution_name and payment_details are required."
            )
            continue
        try:
            Bill.objects.create(
                institution_name=institution_name,
                phone_number=_cell_str(row.get("phone_number")) or None,
                email=_cell_str(row.get("email")) or None,
                address=_cell_str(row.get("address")) or None,
                description=_cell_str(row.get("description")) or None,
                payment_details=payment_details,
                amount=float(_parse_decimal(row.get("amount"), 0)),
                status=_parse_bool(row.get("status")),
            )
            created += 1
        except Exception as exc:
            errors.append(f"Row {i}: {exc}")
    return created, 0, errors


IMPORT_HANDLERS = {
    "inventory": import_inventory_rows,
    "customers": import_customer_rows,
    "suppliers": import_supplier_rows,
    "purchases": import_purchase_rows,
    "sales": import_sale_rows,
    "deliveries": import_delivery_rows,
    "invoices": import_invoice_rows,
    "bills": import_bill_rows,
}
